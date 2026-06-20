# EDM v2 — Export Router (§6.1) — Multi-tenant, Pylon ERP compatible

import csv
import io
from datetime import date, datetime, timezone
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models import Product, ExportLog, User, Supplier
from app.auth import require_role, Role

router = APIRouter(prefix="/api/v1/export", tags=["export"])

# ── Pylon ERP column mapping ──
PYLON_HEADERS = [
    "kwdikos_proiontos",       # ergalyon_code
    "barcode_or_ean",          # ean
    "perigrafi",               # description
    "kwdikos_promithiti",      # supplier_code
    "kwdikos_kataskevasti",    # manufacturer_code
    "timh_cost",               # cost_price (from price_history or current_price)
    "nomisma",                 # price_currency
    "esoterikos_kwdikos",      # internal_sku
    "kwdikos_pylon",           # pylon_code
    "aade_afm_promithiti",     # supplier.afm
    "hmeromhnia_ejagwghs",     # export timestamp
]


async def _build_product_query(
    db: AsyncSession,
    org_id: UUID,
    supplier_id: Optional[UUID] = None,
    category_k1_id: Optional[UUID] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
):
    query = (
        select(Product)
        .where(Product.is_deleted == False)
        .where(Product.organization_id == org_id)
    )
    if supplier_id:
        query = query.where(Product.supplier_id == supplier_id)
    if category_k1_id:
        query = query.where(Product.category_k1_id == category_k1_id)
    if date_from:
        query = query.where(func.date(Product.created_at) >= date_from)
    if date_to:
        query = query.where(func.date(Product.created_at) <= date_to)
    return query


@router.get("")
async def export_products(
    format: str = Query("csv", pattern="^(csv|json|xlsx)$"),
    supplier_id: UUID = Query(None),
    category_k1_id: UUID = Query(None),
    date_from: date = Query(None),
    date_to: date = Query(None),
    pylon_compatible: bool = Query(False, description="Export in Pylon ERP compatible format"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(Role.VIEWER)),
):
    """Export products in CSV, JSON, or XLSX format.
    Pylon-compatible mode uses ERP column names and AFM extraction.
    """
    query = await _build_product_query(
        db, current_user.organization_id, supplier_id, category_k1_id, date_from, date_to
    )
    query = query.order_by(Product.created_at.desc())
    result = await db.execute(query)
    products = result.scalars().all()

    # Pre-load supplier AFM info for Pylon format
    supplier_afm_map = {}
    if pylon_compatible and format in ("csv", "xlsx"):
        suppliers = (await db.execute(
            select(Supplier).where(Supplier.organization_id == current_user.organization_id)
        )).scalars().all()
        supplier_afm_map = {s.id: s.afm for s in suppliers}

    # Log export
    export_log = ExportLog(
        organization_id=current_user.organization_id,
        export_type="products",
        file_format=format,
        status="COMPLETED",
        total_rows=len(products),
        requested_by=current_user.id,
    )
    db.add(export_log)
    await db.flush()

    timestamp_str = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    # ── CSV ──
    if format == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        if pylon_compatible:
            writer.writerow(PYLON_HEADERS)
            for p in products:
                afm = supplier_afm_map.get(p.supplier_id, "")
                writer.writerow([
                    p.ergalyon_code, p.ean, p.description,
                    p.supplier_code, p.manufacturer_code,
                    float(p.current_price) if p.current_price else None,
                    p.price_currency, p.internal_sku, p.pylon_code,
                    afm, timestamp_str,
                ])
        else:
            writer.writerow([
                "ergalyon_code", "supplier_code", "manufacturer_code",
                "ean", "internal_sku", "pylon_code", "description",
                "current_price", "currency",
            ])
            for p in products:
                writer.writerow([
                    p.ergalyon_code, p.supplier_code, p.manufacturer_code,
                    p.ean, p.internal_sku, p.pylon_code, p.description,
                    p.current_price, p.price_currency,
                ])
        output.seek(0)
        filename = f"edm_export_{timestamp_str}.csv"
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    # ── JSON ──
    elif format == "json":
        items = []
        for p in products:
            item = {
                "ergalyon_code": p.ergalyon_code,
                "supplier_code": p.supplier_code,
                "manufacturer_code": p.manufacturer_code,
                "ean": p.ean,
                "internal_sku": p.internal_sku,
                "pylon_code": p.pylon_code,
                "description": p.description,
                "current_price": float(p.current_price) if p.current_price else None,
                "currency": p.price_currency,
            }
            if pylon_compatible:
                item["aade_afm_promithiti"] = supplier_afm_map.get(p.supplier_id, "")
                item["hmeromhnia_ejagwghs"] = timestamp_str
            items.append(item)
        return {"total": len(items), "items": items}

    # ── XLSX ──
    elif format == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        wb = Workbook()

        # Sheet 1: Products
        ws = wb.active
        ws.title = "Products"

        # Styling
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        if pylon_compatible:
            # Pylon ERP format
            ws.append(PYLON_HEADERS)
            for p in products:
                afm = supplier_afm_map.get(p.supplier_id, "")
                ws.append([
                    p.ergalyon_code, p.ean, p.description,
                    p.supplier_code, p.manufacturer_code,
                    float(p.current_price) if p.current_price else None,
                    p.price_currency, p.internal_sku, p.pylon_code,
                    afm, timestamp_str,
                ])
        else:
            headers = [
                "A/A", "ergalyon_code", "supplier_code", "manufacturer_code",
                "ean", "internal_sku", "pylon_code", "description",
                "current_price", "currency", "manufacturer_flag",
                "data_completeness", "created_at",
            ]
            ws.append(headers)
            for idx, p in enumerate(products, 1):
                ws.append([
                    idx,
                    p.ergalyon_code, p.supplier_code, p.manufacturer_code,
                    p.ean, p.internal_sku, p.pylon_code, p.description,
                    float(p.current_price) if p.current_price else None,
                    p.price_currency, "ΝΑΙ" if p.manufacturer_flag else "ΟΧΙ",
                    p.data_completeness_score,
                    p.created_at.isoformat() if p.created_at else "",
                ])

        # Style header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Style data cells
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # Auto-fit column widths
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

        # Freeze header row & auto-filter
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        # Sheet 2: Export Info
        ws2 = wb.create_sheet("Export Info")
        info = [
            ["Ημερομηνία Εξαγωγής", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")],
            ["Πλήθος Προϊόντων", len(products)],
            ["Οργανισμός ID", str(current_user.organization_id)],
            ["Pylon Compatible", "ΝΑΙ" if pylon_compatible else "ΟΧΙ"],
            ["Φίλτρο Supplier", str(supplier_id) if supplier_id else "ΟΛΑ"],
            ["Φίλτρο Κατηγορία", str(category_k1_id) if category_k1_id else "ΟΛΕΣ"],
        ]
        for row in info:
            ws2.append(row)
        ws2.column_dimensions["A"].width = 30
        ws2.column_dimensions["B"].width = 50

        # Save
        virtual_workbook = BytesIO()
        wb.save(virtual_workbook)
        virtual_workbook.seek(0)

        filename = f"edm_export_{timestamp_str}.xlsx"
        return StreamingResponse(
            virtual_workbook,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    else:
        raise HTTPException(status_code=400, detail=f"Format '{format}' not yet implemented")